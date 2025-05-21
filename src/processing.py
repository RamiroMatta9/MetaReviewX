import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from translations import TranslationManager

translation_manager = TranslationManager()

def verificar_archivos(archivos, language):
    """Verifica si los archivos existen y son legibles"""
    archivos_validos = []
    problemas = []
    
    for archivo in archivos:
        if not os.path.exists(archivo):
            problemas.append(translation_manager.get_translation(language, 'file_not_found', archivo))
        elif not archivo.lower().endswith('.csv'):
            problemas.append(translation_manager.get_translation(language, 'invalid_csv', archivo))
        else:
            archivos_validos.append(archivo)
    
    return archivos_validos, problemas

def extraer_año(fecha_str, columna_fecha='Publication Year'):
    """Intenta extraer el año de diferentes formatos de fecha"""
    if pd.isna(fecha_str):
        return None
    
    # Si ya es un número (año directo)
    try:
        return int(float(fecha_str))
    except:
        pass
    
    # Para strings con formato de fecha
    try:
        return pd.to_datetime(fecha_str, errors='coerce').year
    except:
        return None

def procesar_archivos(archivos, language):
    """Procesa y unifica los archivos CSV"""
    dataframes = []
    for archivo in archivos:
        try:
            df = pd.read_csv(archivo, delimiter=None, engine='python', encoding='utf-8')
            df['Fuente'] = os.path.splitext(os.path.basename(archivo))[0]
            dataframes.append(df)
        except Exception as e:
            print(translation_manager.get_translation(language, 'processing_error', archivo, str(e)))
    
    if not dataframes:
        return None
    
    return pd.concat(dataframes, ignore_index=True)

def identificar_duplicados(df, language, columna_titulo='Title'):
    """Identifica duplicados y genera reporte"""
    report_lines = []
    
    if columna_titulo not in df.columns:
        report_lines.append(translation_manager.get_translation(language, 'no_title_column'))
        return df, [], report_lines
    
    duplicados = df[df.duplicated(subset=[columna_titulo], keep=False)]
    conteo_duplicados = duplicados[columna_titulo].value_counts()
    
    if not duplicados.empty:
        report_lines.append("\n" + "="*60)
        report_lines.append(translation_manager.get_translation(language, 'duplicates_report'))
        report_lines.append("="*60)
        report_lines.append(translation_manager.get_translation(language, 'total_duplicates', len(duplicados)))
        report_lines.append(translation_manager.get_translation(language, 'unique_duplicates', len(conteo_duplicados)))
        report_lines.append("\n" + translation_manager.get_translation(language, 'top_duplicates'))
        report_lines.append(str(conteo_duplicados.head(10)))
    
    return df, duplicados.index.tolist(), report_lines

def crear_excel_organizado(df, archivo_salida, indices_duplicados, language, columna_fecha='Publication Year'):
    """Crea un Excel con organización por años"""
    wb = Workbook()
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Extraer años
    df['Año_Extraido'] = df[columna_fecha].apply(extraer_año)
    años_validos = df['Año_Extraido'].dropna().unique()
    
    if len(años_validos) == 0:
        print("\n" + translation_manager.get_translation(language, 'no_valid_years'))
        años_validos = ['Todos']  # Hoja por defecto
    
    # 1. Hoja principal "Todos" con duplicados marcados
    ws_todos = wb.create_sheet(translation_manager.get_translation(language, 'all_sheet'))
    for r_idx, row in enumerate(dataframe_to_rows(df.drop('Año_Extraido', axis=1), index=False, header=True), 1):
        ws_todos.append(row)
        if r_idx > 1 and (r_idx-2) in indices_duplicados:
            for cell in ws_todos[r_idx]:
                cell.fill = fill
    
    # 2. Hojas por año (sin marcar duplicados)
    for año in sorted(años_validos):
        if str(año) == 'Todos':
            continue  # Ya creamos esta hoja
        
        df_año = df[df['Año_Extraido'] == año]
        nombre_hoja = f"{translation_manager.get_translation(language, 'year')} {int(año)}"[:31]
        ws = wb.create_sheet(nombre_hoja)
        
        for row in dataframe_to_rows(df_año.drop('Año_Extraido', axis=1), index=False, header=True):
            ws.append(row)
    
    # 3. Hoja de resumen
    ws_resumen = wb.create_sheet(translation_manager.get_translation(language, 'summary_sheet'))
    resumen_data = {
        translation_manager.get_translation(language, 'year'): [],
        translation_manager.get_translation(language, 'total_articles'): [],
        translation_manager.get_translation(language, 'unique_articles'): [],
        translation_manager.get_translation(language, 'duplicate_count'): [],
        translation_manager.get_translation(language, 'sources'): []
    }
    
    # Datos para todos los años
    resumen_data[translation_manager.get_translation(language, 'year')].append('Todos')
    resumen_data[translation_manager.get_translation(language, 'total_articles')].append(len(df))
    resumen_data[translation_manager.get_translation(language, 'unique_articles')].append(len(df['Title'].unique()) if 'Title' in df.columns else 'N/A')
    resumen_data[translation_manager.get_translation(language, 'duplicate_count')].append(len(indices_duplicados))
    resumen_data[translation_manager.get_translation(language, 'sources')].append(", ".join(df['Fuente'].unique()))
    
    # Datos por año
    for año in sorted(años_validos):
        if str(año) == 'Todos':
            continue
        
        df_año = df[df['Año_Extraido'] == año]
        resumen_data[translation_manager.get_translation(language, 'year')].append(str(int(año)))
        resumen_data[translation_manager.get_translation(language, 'total_articles')].append(len(df_año))
        resumen_data[translation_manager.get_translation(language, 'unique_articles')].append(len(df_año['Title'].unique()) if 'Title' in df_año.columns else 'N/A')
        resumen_data[translation_manager.get_translation(language, 'duplicate_count')].append(sum(1 for i in df_año.index if i in indices_duplicados))
        resumen_data[translation_manager.get_translation(language, 'sources')].append(", ".join(df_año['Fuente'].unique()))
    
    # Escribir resumen
    for col_idx, (header, data) in enumerate(resumen_data.items(), start=1):
        ws_resumen.cell(row=1, column=col_idx, value=header)
        for row_idx, value in enumerate(data, start=2):
            ws_resumen.cell(row=row_idx, column=col_idx, value=value)
    
    # Eliminar hoja vacía inicial y ajustar columnas
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2
    
    # Guardar archivo
    wb.save(archivo_salida)

def process_files(input_files, output_file, language='es'):
    """Función principal para procesar archivos"""
    report_lines = []
    
    report_lines.append("\n" + "="*60)
    report_lines.append(translation_manager.get_translation(language, 'window_title'))
    report_lines.append("="*60)
    
    # Verificar archivos
    archivos_validos, problemas = verificar_archivos(input_files, language)
    
    if problemas:
        report_lines.append("\n" + translation_manager.get_translation(language, 'report_title').split(':')[0] + ":")
        for problema in problemas:
            report_lines.append(f" - {problema}")
    
    if not archivos_validos:
        report_lines.append("\n" + translation_manager.get_translation(language, 'no_valid_files'))
        return False, "\n".join(report_lines)
    
    # Procesar datos
    df = procesar_archivos(archivos_validos, language)
    if df is None:
        report_lines.append("\n" + translation_manager.get_translation(language, 'no_valid_files'))
        return False, "\n".join(report_lines)
    
    # Identificar duplicados
    df, dup_indices, dup_report = identificar_duplicados(df, language)
    report_lines.extend(dup_report)
    
    # Crear Excel organizado
    report_lines.append("\n" + translation_manager.get_translation(language, 'process_completed'))
    crear_excel_organizado(df, output_file, dup_indices, language)
    
    report_lines.append("\n" + "="*60)
    report_lines.append(translation_manager.get_translation(language, 'process_completed'))
    report_lines.append("="*60)
    report_lines.append("\n" + translation_manager.get_translation(language, 'output_file', output_file))
    report_lines.append(translation_manager.get_translation(language, 'total_records', len(df)))
    report_lines.append(translation_manager.get_translation(language, 'duplicates', len(dup_indices)))
    report_lines.append("\n" + translation_manager.get_translation(language, 'file_contents'))
    report_lines.append(translation_manager.get_translation(language, 'all_sheet'))
    report_lines.append(translation_manager.get_translation(language, 'year_sheets'))
    report_lines.append(translation_manager.get_translation(language, 'summary_sheet'))
    report_lines.append("\n" + translation_manager.get_translation(language, 'footer'))
    
    return True, "\n".join(report_lines)